import zipfile
from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook

from app.verticals.real_estate.template_filling.excel.template_filler import (
    TemplateFiller,
    WORKSHEET_XML_NS,
    X14_NS,
    _extract_dropdown_extlsts,
    _restore_dropdown_extlsts,
)

lxml_etree = pytest.importorskip("lxml.etree")


def _worksheet_xml_with_dropdown() -> bytes:
    return f"""<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<worksheet xmlns="{WORKSHEET_XML_NS}"
 xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships"
 xmlns:mc="http://schemas.openxmlformats.org/markup-compatibility/2006"
 xmlns:x14="{X14_NS}"
 xmlns:xm="http://schemas.microsoft.com/office/excel/2006/main"
 xmlns:xr="http://schemas.microsoft.com/office/spreadsheetml/2014/revision"
 mc:Ignorable="x14 xr">
  <sheetData>
    <row r="1">
      <c r="A1" t="inlineStr">
        <is><t>OFF</t></is>
      </c>
    </row>
  </sheetData>
  <extLst>
    <ext uri="{{CCE6A557-97BC-4B89-ADB6-D9C93CAAB3DF}}">
      <x14:dataValidations count="1">
        <x14:dataValidation type="list" allowBlank="1" showInputMessage="1" showErrorMessage="1" xr:uid="{{12345678-1234-1234-1234-123456789ABC}}">
          <x14:formula1>"OFF,ON"</x14:formula1>
          <xm:sqref>A1</xm:sqref>
        </x14:dataValidation>
      </x14:dataValidations>
    </ext>
  </extLst>
</worksheet>
""".encode("utf-8")


def _worksheet_xml_without_dropdown() -> bytes:
    return f"""<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<worksheet xmlns="{WORKSHEET_XML_NS}">
  <sheetData>
    <row r="1">
      <c r="A1" t="inlineStr">
        <is><t>OFF</t></is>
      </c>
    </row>
  </sheetData>
</worksheet>
""".encode("utf-8")


def _write_zip_with_sheet(path: Path, sheet_xml: bytes, sheet_name: str = "xl/worksheets/sheet1.xml") -> None:
    with zipfile.ZipFile(path, "w", compression=zipfile.ZIP_DEFLATED) as workbook_zip:
        workbook_zip.writestr("[Content_Types].xml", "<Types/>")
        workbook_zip.writestr(sheet_name, sheet_xml)


def _replace_zip_entry(path: Path, entry_name: str, entry_bytes: bytes) -> None:
    temp_path = path.with_suffix(path.suffix + ".patched")

    with zipfile.ZipFile(path, "r") as source_zip:
        with zipfile.ZipFile(temp_path, "w", compression=zipfile.ZIP_DEFLATED) as patched_zip:
            for info in source_zip.infolist():
                if info.filename == entry_name:
                    patched_zip.writestr(info, entry_bytes)
                else:
                    patched_zip.writestr(info, source_zip.read(info.filename))

    temp_path.replace(path)


def test_extract_dropdown_extlsts_preserves_used_namespaces(tmp_path):
    workbook_path = tmp_path / "template.xlsx"
    _write_zip_with_sheet(workbook_path, _worksheet_xml_with_dropdown(), sheet_name="xl/worksheets/sheet3.xml")

    preserved = _extract_dropdown_extlsts(str(workbook_path))

    assert "xl/worksheets/sheet3.xml" in preserved
    captured = preserved["xl/worksheets/sheet3.xml"]

    assert captured.namespaces == {
        "x14": X14_NS,
        "xm": "http://schemas.microsoft.com/office/excel/2006/main",
        "xr": "http://schemas.microsoft.com/office/spreadsheetml/2014/revision",
    }
    assert b"x14:dataValidations" in captured.extlst_xml
    lxml_etree.fromstring(captured.extlst_xml)


def test_restore_dropdown_extlsts_adds_root_namespaces_and_single_extlst(tmp_path):
    template_path = tmp_path / "template.xlsx"
    output_path = tmp_path / "output.xlsx"

    _write_zip_with_sheet(template_path, _worksheet_xml_with_dropdown())
    _write_zip_with_sheet(output_path, _worksheet_xml_without_dropdown())

    preserved = _extract_dropdown_extlsts(str(template_path))
    _restore_dropdown_extlsts(str(output_path), preserved)

    with zipfile.ZipFile(output_path, "r") as workbook_zip:
        patched_xml = workbook_zip.read("xl/worksheets/sheet1.xml")

    root = lxml_etree.fromstring(patched_xml)
    assert root.nsmap["x14"] == X14_NS
    assert root.nsmap["xm"] == "http://schemas.microsoft.com/office/excel/2006/main"
    assert root.nsmap["xr"] == "http://schemas.microsoft.com/office/spreadsheetml/2014/revision"

    extlst_nodes = root.findall(f"{{{WORKSHEET_XML_NS}}}extLst")
    assert len(extlst_nodes) == 1
    assert root.xpath("count(//x14:dataValidations)", namespaces={"x14": X14_NS}) == 1.0
    lxml_etree.fromstring(patched_xml)


def test_fill_template_restores_dropdown_metadata_and_output_reopens(tmp_path):
    template_path = tmp_path / "template.xlsx"
    output_path = tmp_path / "filled.xlsx"

    workbook = Workbook()
    workbook.active.title = "Sheet1"
    workbook.active["A1"] = "OFF"
    workbook.save(template_path)
    workbook.close()

    _replace_zip_entry(template_path, "xl/worksheets/sheet1.xml", _worksheet_xml_with_dropdown())

    filler = TemplateFiller()
    summary = filler.fill_template(
        str(template_path),
        str(output_path),
        field_mapping={"mappings": []},
        extracted_data={},
    )

    assert summary["errors"] == []

    with zipfile.ZipFile(output_path, "r") as workbook_zip:
        patched_xml = workbook_zip.read("xl/worksheets/sheet1.xml")

    root = lxml_etree.fromstring(patched_xml)
    assert root.find(f"{{{WORKSHEET_XML_NS}}}extLst") is not None
    assert root.xpath("count(//x14:dataValidation)", namespaces={"x14": X14_NS}) == 1.0

    reopened = load_workbook(output_path, data_only=False)
    reopened.close()


def test_fill_template_uses_keep_vba_for_xlsm(tmp_path, monkeypatch):
    template_path = tmp_path / "template.xlsm"
    output_path = tmp_path / "filled.xlsm"
    _write_zip_with_sheet(template_path, _worksheet_xml_without_dropdown())

    calls = {}

    class FakeWorkbook:
        sheetnames = []
        defined_names = {}
        _external_links = []

        def save(self, _output_path):
            return None

        def close(self):
            return None

    def fake_load_workbook(path, data_only=False, keep_vba=False):
        calls["path"] = path
        calls["data_only"] = data_only
        calls["keep_vba"] = keep_vba
        return FakeWorkbook()

    monkeypatch.setattr("openpyxl.load_workbook", fake_load_workbook)

    filler = TemplateFiller()
    filler.fill_template(
        str(template_path),
        str(output_path),
        field_mapping={"mappings": []},
        extracted_data={},
    )

    assert calls["path"] == str(template_path)
    assert calls["data_only"] is False
    assert calls["keep_vba"] is True


def test_template_without_dropdown_extlst_saves_normally(tmp_path):
    template_path = tmp_path / "template.xlsx"
    output_path = tmp_path / "output.xlsx"

    workbook = Workbook()
    workbook.active.title = "Sheet1"
    workbook.active["A1"] = "plain"
    workbook.save(template_path)
    workbook.close()

    filler = TemplateFiller()
    summary = filler.fill_template(
        str(template_path),
        str(output_path),
        field_mapping={"mappings": []},
        extracted_data={},
    )

    assert summary["errors"] == []
    reopened = load_workbook(output_path, data_only=False)
    assert "Sheet1" in reopened.sheetnames
    reopened.close()
