"""Cell and sheet styling inspection utilities for Excel templates."""

from typing import Optional
from openpyxl.cell import Cell
from openpyxl.worksheet.worksheet import Worksheet

from app.utils.logging import logger


class StyleInspector:
    """Inspect cell and sheet styling for table detection and analysis."""

    def _has_border(self, cell: Cell, side: str = 'any') -> bool:
        """
        Check if a cell has a border on a specific side.

        Args:
            cell: Cell to check
            side: 'top', 'bottom', 'left', 'right', or 'any'

        Returns:
            True if cell has border on specified side
        """
        if not cell.border:
            return False

        border_styles = {
            'top': cell.border.top and cell.border.top.style,
            'bottom': cell.border.bottom and cell.border.bottom.style,
            'left': cell.border.left and cell.border.left.style,
            'right': cell.border.right and cell.border.right.style,
        }

        if side == 'any':
            return any(border_styles.values())
        else:
            return bool(border_styles.get(side))

    def _has_fill(self, cell: Cell) -> bool:
        """Return True if a cell likely has a non-default fill/background."""
        try:
            fill = getattr(cell, 'fill', None)
            if not fill:
                return False

            pattern_type = getattr(fill, 'patternType', None)
            if pattern_type and str(pattern_type).lower() != 'none':
                fg = getattr(fill, 'fgColor', None) or getattr(fill, 'start_color', None)
                rgb = getattr(fg, 'rgb', None)
                if rgb and str(rgb).upper() in {'00000000', 'FFFFFFFF'}:
                    return False
                return True

            fg = getattr(fill, 'fgColor', None) or getattr(fill, 'start_color', None)
            rgb = getattr(fg, 'rgb', None)
            if rgb and str(rgb).upper() not in {'00000000', 'FFFFFFFF'}:
                return True
        except Exception:
            return False

        return False

    def _is_bold(self, cell: Cell) -> bool:
        """Check if cell has bold font."""
        try:
            font = getattr(cell, 'font', None)
            if not font:
                return False
            return getattr(font, 'bold', False) or False
        except Exception:
            return False

    def _get_font_size(self, cell: Cell) -> Optional[float]:
        """Get cell font size (None if default/unspecified)."""
        try:
            font = getattr(cell, 'font', None)
            if not font:
                return None
            return getattr(font, 'size', None)
        except Exception:
            return None

    def _get_font_color(self, cell: Cell) -> Optional[str]:
        """Get cell font color as RGB string (None if default)."""
        try:
            font = getattr(cell, 'font', None)
            if not font:
                return None
            color = getattr(font, 'color', None)
            if not color:
                return None
            rgb = getattr(color, 'rgb', None)
            if rgb:
                return str(rgb).upper()
            return None
        except Exception:
            return None

    def _get_fill_color(self, cell: Cell) -> Optional[str]:
        """Get cell fill/background color as RGB string (None if default)."""
        try:
            fill = getattr(cell, 'fill', None)
            if not fill:
                return None

            # Try foreground color
            fg = getattr(fill, 'fgColor', None) or getattr(fill, 'start_color', None)
            if fg:
                rgb = getattr(fg, 'rgb', None)
                if rgb:
                    rgb_str = str(rgb).upper()
                    # Filter out default colors (transparent, white)
                    if rgb_str not in {'00000000', 'FFFFFFFF'}:
                        return rgb_str
            return None
        except Exception:
            return None

    def _get_alignment(self, cell: Cell) -> Optional[str]:
        """Get horizontal alignment (left, center, right, justify)."""
        try:
            alignment = getattr(cell, 'alignment', None)
            if not alignment:
                return None
            return getattr(alignment, 'horizontal', None)
        except Exception:
            return None

    def _has_data_validation(self, sheet: Worksheet, row: int, col: int) -> bool:
        """Check if cell has data validation (dropdown, list, etc.)."""
        try:
            if not hasattr(sheet, 'data_validations'):
                return False

            for dv in sheet.data_validations.dataValidation:
                if not hasattr(dv, 'cells'):
                    continue

                # Check if cell is in any validation range
                for cell_range in dv.cells.ranges:
                    if hasattr(cell_range, 'min_row'):
                        if (cell_range.min_row <= row <= cell_range.max_row and
                            cell_range.min_col <= col <= cell_range.max_col):
                            return True
            return False
        except Exception:
            return False

    def _get_frozen_rows(self, sheet: Worksheet) -> int:
        """
        Get number of frozen rows (often indicates header rows).

        Returns:
            Number of rows that are frozen (0 if none)
        """
        try:
            if not sheet.freeze_panes:
                return 0

            # freeze_panes is like "A3" meaning rows 1-2 are frozen
            frozen_cell = sheet.freeze_panes
            if frozen_cell:
                # Get row number from coordinate (e.g., "A3" -> row 3)
                # Rows above this are frozen
                return frozen_cell.row - 1
            return 0
        except Exception:
            return 0

    def _row_has_consistent_styling(
        self,
        sheet: Worksheet,
        row_idx: int,
        start_col: int,
        end_col: int
    ) -> bool:
        """
        Check if entire row has consistent styling (likely header).

        Checks if all cells in the row have same fill color and bold status.
        """
        try:
            if end_col < start_col:
                return False

            first_cell = sheet.cell(row_idx, start_col)
            first_fill = self._get_fill_color(first_cell)
            first_bold = self._is_bold(first_cell)

            # If first cell has no styling, not a styled header
            if not first_fill and not first_bold:
                return False

            # Check remaining cells for consistency
            for col_idx in range(start_col + 1, end_col + 1):
                cell = sheet.cell(row_idx, col_idx)
                if self._get_fill_color(cell) != first_fill:
                    return False
                if self._is_bold(cell) != first_bold:
                    return False

            return True
        except Exception:
            return False

    def _row_has_table_markers(self, sheet: Worksheet, row_idx: int, start_col: int, end_col: int) -> bool:
        """Heuristic: does this row look like part of a table (borders/fill)?"""
        for col_idx in range(start_col, end_col + 1):
            cell = sheet.cell(row_idx, col_idx)
            if self._has_border(cell, 'any') or self._has_fill(cell):
                return True
        return False
