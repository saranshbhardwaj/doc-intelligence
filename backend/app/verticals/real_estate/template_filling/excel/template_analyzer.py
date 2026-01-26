"""Template analysis orchestration for detecting fields and tables."""

from typing import Any, Dict, List, Optional, Tuple
from openpyxl.cell import Cell
from openpyxl.utils.cell import range_boundaries
from openpyxl.worksheet.worksheet import Worksheet

from app.utils.logging import logger
from .table_detector import TableDetector
from .style_inspector import StyleInspector


class TemplateAnalyzer:
    """Analyze Excel templates to detect fillable cells, tables, and structure."""

    def __init__(self):
        """Initialize template analyzer with dependencies."""
        self._table_detector = TableDetector()
        self._style_inspector = StyleInspector()

    def analyze_template(self, workbook) -> Dict[str, Any]:
        """
        Analyze template structure (called by facade).

        Args:
            workbook: openpyxl Workbook object

        Returns:
            Dictionary with complete template schema
        """
        sheets_data = []
        total_kv_fields = 0
        total_tables = 0

        for sheet_idx, sheet_name in enumerate(workbook.sheetnames):
            sheet = workbook[sheet_name]

            # Detect key-value pairs
            kv_fields = self._detect_key_value_fields(sheet)

            # Detect table structures with merged cell support
            tables = self._detect_tables(sheet)

            # Detect formula cells
            formula_cells = self._detect_formula_cells(sheet)

            total_kv_fields += len(kv_fields)
            total_tables += len(tables)

            sheet_data = {
                "name": sheet_name,
                "index": sheet_idx,
                "key_value_fields": kv_fields,
                "tables": tables,
                "formula_cells": formula_cells,
                "total_rows": sheet.max_row,
                "total_cols": sheet.max_column,
            }

            sheets_data.append(sheet_data)

        schema = {
            "sheets": sheets_data,
            "total_key_value_fields": total_kv_fields,
            "total_tables": total_tables,
            "has_formulas": any(len(s["formula_cells"]) > 0 for s in sheets_data),
        }

        logger.info(
            f"Template analysis complete: {len(sheets_data)} sheets, "
            f"{total_kv_fields} key-value fields, {total_tables} tables"
        )

        return schema

    def _detect_tables(self, sheet: Worksheet) -> List[Dict[str, Any]]:
        """
        Detect tables with merged cell support.

        Wraps TableDetector.detect_tables() to inject merged cell handling.

        Args:
            sheet: Worksheet

        Returns:
            List of detected tables
        """
        # Save original method
        original_analyze = self._table_detector._analyze_table_structure

        # Override with our enhanced version (accepting both positional and keyword args)
        def override_analyze(sh, hr, hc, end_row=None):
            return self._analyze_table_structure_impl(sh, hr, hc, end_row)

        self._table_detector._analyze_table_structure = override_analyze

        # Detect tables
        tables = self._table_detector.detect_tables(sheet)

        # Restore original method
        self._table_detector._analyze_table_structure = original_analyze

        return tables

    def _analyze_table_structure_impl(
        self,
        sheet: Worksheet,
        header_row: int,
        header_cells: List[tuple],
        end_row: Optional[int] = None,
    ) -> Optional[Dict[str, Any]]:
        """
        Analyze table structure with merged cell support.

        Args:
            sheet: Worksheet
            header_row: Row index of headers
            header_cells: List of (col_idx, header_text) tuples
            end_row: Optional end row of table data

        Returns:
            Table metadata dict or None
        """
        from openpyxl.utils import get_column_letter

        start_col = min(c[0] for c in header_cells)
        end_col = max(c[0] for c in header_cells)

        # Detect multi-level header rows with merged cell support
        header_rows = self._detect_header_rows_with_merged(sheet, header_row, start_col, end_col)

        # Build hierarchical column headers from all detected header levels
        column_headers = self._build_hierarchical_headers_impl(sheet, header_rows, start_col, end_col)

        # Log hierarchical header detection
        if len(header_rows) > 1:
            logger.info(
                f"Table at row {header_row} in '{sheet.title}': detected {len(header_rows)} header levels "
                f"(rows {header_rows}), {len(column_headers)} columns with hierarchical names"
            )
            sample_headers = [h["header"] for h in column_headers[:5]]
            logger.info(f"  Sample hierarchical headers: {sample_headers}")

        # Check for row header column
        row_header_col = None
        if start_col > 1:
            left_cell = sheet.cell(header_row, start_col - 1)
            if left_cell.value and left_cell.data_type != 'f':
                row_header_col = start_col - 1

        # Find data rows below headers
        data_rows = []
        fillable_cells = []
        max_data_rows = 50

        if end_row is not None:
            end_row = min(end_row, sheet.max_row)
            if end_row <= header_row:
                return None
            max_data_rows = min(max_data_rows, end_row - header_row)

        for row_offset in range(1, max_data_rows + 1):
            data_row_idx = header_row + row_offset
            if data_row_idx > sheet.max_row:
                break

            if end_row is not None and data_row_idx > end_row:
                break

            row_has_content = False
            row_data = []
            row_has_markers = self._style_inspector._row_has_table_markers(sheet, data_row_idx, start_col, end_col)

            row_label = None
            if row_header_col:
                row_header_cell = sheet.cell(data_row_idx, row_header_col)
                if row_header_cell.value and row_header_cell.data_type != 'f':
                    row_label = str(row_header_cell.value).strip()
                    row_has_content = True

            for col_idx in range(start_col, end_col + 1):
                cell = sheet.cell(data_row_idx, col_idx)
                cell_has_marker = self._style_inspector._has_border(cell, 'any') or self._style_inspector._has_fill(cell)

                if cell.data_type != 'f':
                    is_empty = cell.value is None or str(cell.value).strip() == ""

                    if is_empty:
                        if end_row is None and not row_has_markers and not cell_has_marker and not row_label:
                            row_data.append(cell.value)
                            continue

                        col_header_info = next((h for h in column_headers if h["col"] == col_idx), None)
                        col_header = col_header_info["header"] if col_header_info else ""
                        col_leaf_header = col_header_info["leaf_header"] if col_header_info else ""

                        fillable_cells.append({
                            "cell": cell.coordinate,
                            "row": data_row_idx,
                            "col": col_idx,
                            "col_letter": get_column_letter(col_idx),
                            "row_label": row_label,
                            "col_header": col_header,
                            "col_leaf_header": col_leaf_header,
                            "type": self._infer_cell_type(cell),
                        })
                        row_has_content = True
                    else:
                        row_has_content = True

                if cell_has_marker:
                    row_has_content = True

                row_data.append(cell.value)

            if not row_has_content:
                break

            data_rows.append(data_row_idx)

        if len(data_rows) == 0:
            return None

        table_name = None
        topmost_header_row = header_rows[0] if header_rows else header_row
        if topmost_header_row > 1:
            title_row = topmost_header_row - 1
            title_cell = sheet.cell(title_row, start_col)
            if title_cell.value and title_cell.data_type != 'f':
                table_name = str(title_cell.value).strip()

        return {
            "table_name": table_name or f"Table at row {header_row}",
            "start_row": topmost_header_row,
            "start_col": start_col,
            "end_col": end_col,
            "end_row": (end_row if end_row is not None else max(data_rows)),
            "header_row": header_row,
            "header_rows": header_rows,
            "row_header_col": row_header_col,
            "column_headers": [h["header"] for h in column_headers],
            "column_headers_detailed": column_headers,
            "data_rows": data_rows,
            "total_data_rows": len(data_rows),
            "total_fillable_cells": len(fillable_cells),
            "fillable_cells": fillable_cells[:100],
        }

    def _detect_header_rows_with_merged(
        self,
        sheet: Worksheet,
        potential_header_row: int,
        start_col: int,
        end_col: int
    ) -> List[int]:
        """
        Detect multiple header rows with merged cell support.

        Args:
            sheet: Worksheet
            potential_header_row: The row index that looks like the main header
            start_col: Starting column
            end_col: Ending column

        Returns:
            List of header row indices in order (top to bottom)
        """
        header_rows = [potential_header_row]

        for offset in range(1, 2):
            check_row = potential_header_row - offset
            if check_row < 1:
                break

            has_parent_headers = False
            for col_idx in range(start_col, end_col + 1):
                cell_value = self._get_merged_cell_value(sheet, check_row, col_idx)
                if cell_value and str(cell_value).strip():
                    merge_range = self._get_merged_cell_range(sheet, check_row, col_idx)
                    if merge_range:
                        _, min_col, _, max_col = merge_range
                        if max_col - min_col >= 1:
                            has_parent_headers = True
                            break

            if has_parent_headers:
                header_rows.insert(0, check_row)
            else:
                break

        return header_rows

    def _build_hierarchical_headers_impl(
        self,
        sheet: Worksheet,
        header_rows: List[int],
        start_col: int,
        end_col: int
    ) -> List[Dict[str, Any]]:
        """
        Build hierarchical headers with merged cell support.

        Args:
            sheet: Worksheet
            header_rows: List of header row indices
            start_col: Starting column
            end_col: Ending column

        Returns:
            List of column header dictionaries
        """
        from openpyxl.utils import get_column_letter

        column_headers = []

        for col_idx in range(start_col, end_col + 1):
            header_parts = []

            for header_row_idx in header_rows:
                cell_value = self._get_merged_cell_value(sheet, header_row_idx, col_idx)
                if cell_value:
                    header_text = str(cell_value).strip()
                    if header_text and (not header_parts or header_parts[-1] != header_text):
                        header_parts.append(header_text)

            if header_parts:
                hierarchical_name = " | ".join(header_parts)
                leaf_name = header_parts[-1] if header_parts else ""
            else:
                hierarchical_name = ""
                leaf_name = ""

            column_headers.append({
                "col": col_idx,
                "col_letter": get_column_letter(col_idx),
                "header": hierarchical_name,
                "leaf_header": leaf_name,
                "header_parts": header_parts,
            })

        return column_headers

    def _detect_key_value_fields(self, sheet: Worksheet) -> List[Dict[str, Any]]:
        """
        Detect key-value pair fields in a worksheet.

        Args:
            sheet: openpyxl Worksheet object

        Returns:
            List of key-value field metadata
        """
        kv_fields = []

        # Iterate through reasonable range (avoid scanning entire empty grid)
        max_scan_row = min(sheet.max_row, 200)
        max_scan_col = min(sheet.max_column, 50)

        for row_idx in range(1, max_scan_row + 1):
            for col_idx in range(1, max_scan_col + 1):
                cell = sheet.cell(row_idx, col_idx)

                if self._is_fillable_cell(cell, sheet):
                    # Try to find a meaningful label
                    label = self._find_cell_label(cell, sheet)

                    # Only include if we found a meaningful label (filters noise)
                    if label and len(label) > 2 and any(c.isalpha() for c in label):
                        kv_fields.append({
                            "cell": cell.coordinate,
                            "label": label,
                            "row": cell.row,
                            "col": cell.column,
                            "type": self._infer_cell_type(cell),
                            "current_value": self._get_cell_display_value(cell),
                            "is_merged": isinstance(cell, Cell) and cell.coordinate in sheet.merged_cells,
                        })

        return kv_fields

    def _is_fillable_cell(self, cell: Cell, sheet: Worksheet) -> bool:
        """
        Determine if a cell is fillable.

        Args:
            cell: openpyxl Cell object
            sheet: Parent worksheet

        Returns:
            True if cell appears to be fillable
        """
        # Skip if cell has a formula
        if cell.data_type == 'f':
            return False

        # Empty cells are potentially fillable
        if cell.value is None or str(cell.value).strip() == "":
            # But only if they're in a reasonable range (not way out in the grid)
            if cell.row <= 200 and cell.column <= 50:
                return True

        # Cells with placeholder patterns
        if cell.value:
            val_str = str(cell.value).strip()
            placeholder_patterns = [
                "[",  # [Name], [Value]
                "enter",  # Enter value
                "input",  # Input here
                "xxx",  # XXX placeholder
                "___",  # Underscores
                "...",  # Ellipsis
            ]
            val_lower = val_str.lower()
            if any(pattern in val_lower for pattern in placeholder_patterns):
                return True

        return False

    def _find_cell_label(self, cell: Cell, sheet: Worksheet) -> Optional[str]:
        """
        Find a label for a fillable cell by looking at nearby cells.

        Strategy:
        1. Check cell directly to the left (immediate label)
        2. Check cell directly above
        3. Check diagonal (top-left)
        4. Search for section headers above (bold/styled cells)

        Args:
            cell: Target cell
            sheet: Parent worksheet

        Returns:
            Label text (may include section prefix like "DEBT | Rate")
        """
        immediate_label = None

        # Check left
        if cell.column > 1:
            left_cell = sheet.cell(row=cell.row, column=cell.column - 1)
            if left_cell.value and left_cell.data_type != 'f':
                label_text = str(left_cell.value).strip()
                if label_text and len(label_text) < 100:  # Reasonable label length
                    immediate_label = label_text

        # Check above (if no left label)
        if not immediate_label and cell.row > 1:
            above_cell = sheet.cell(row=cell.row - 1, column=cell.column)
            if above_cell.value and above_cell.data_type != 'f':
                label_text = str(above_cell.value).strip()
                if label_text and len(label_text) < 100:
                    immediate_label = label_text

        # Check diagonal (top-left) (if still no label)
        if not immediate_label and cell.row > 1 and cell.column > 1:
            diag_cell = sheet.cell(row=cell.row - 1, column=cell.column - 1)
            if diag_cell.value and diag_cell.data_type != 'f':
                label_text = str(diag_cell.value).strip()
                if label_text and len(label_text) < 100:
                    immediate_label = label_text

        # Search for section header above (bold/styled text that could be a section title)
        section_header = self._find_section_header(cell, sheet)

        # Combine section header with immediate label
        if section_header and immediate_label:
            # Only add section prefix if it's different from the immediate label
            if section_header.upper() != immediate_label.upper():
                return f"{section_header} | {immediate_label}"

        return immediate_label

    def _find_section_header(self, cell: Cell, sheet: Worksheet) -> Optional[str]:
        """
        Search for a section header above the cell.

        Looks for bold/styled cells that appear to be section titles.
        Section headers are typically:
        - Bold text
        - Located 1-10 rows above the cell
        - In the same column or slightly to the left

        Args:
            cell: Target cell
            sheet: Parent worksheet

        Returns:
            Section header text or None
        """
        # Search up to 10 rows above for a section header
        max_search_rows = min(cell.row - 1, 10)

        for row_offset in range(1, max_search_rows + 1):
            check_row = cell.row - row_offset

            # Check cells from the cell's column to a few columns left
            # Section headers often span or are positioned at the start of a group
            for col_offset in range(0, min(cell.column, 5)):
                check_col = cell.column - col_offset
                if check_col < 1:
                    continue

                check_cell = sheet.cell(row=check_row, column=check_col)

                # Skip empty cells or formulas
                if not check_cell.value or check_cell.data_type == 'f':
                    continue

                # Check if cell appears to be a header (bold, short text, all caps common)
                is_header = False

                # Check for bold font
                if check_cell.font and check_cell.font.bold:
                    is_header = True

                # Check for merged cell (headers often span multiple columns)
                merge_range = self._get_merged_cell_range(sheet, check_row, check_col)
                if merge_range:
                    _, min_col, _, max_col = merge_range
                    if max_col - min_col >= 1:  # Spans at least 2 columns
                        is_header = True

                if is_header:
                    header_text = str(check_cell.value).strip()
                    # Valid header: short, has alphabetic chars, not a number
                    if (header_text and
                        len(header_text) < 50 and
                        any(c.isalpha() for c in header_text) and
                        not header_text.replace('.', '').replace(',', '').replace('-', '').isdigit()):
                        return header_text

        return None

    def _get_merged_cell_range(self, sheet: Worksheet, row: int, col: int) -> Optional[Tuple[int, int, int, int]]:
        """
        Get the boundaries of a merged cell range if cell is part of one.

        Args:
            sheet: Worksheet
            row: Row index (1-based)
            col: Column index (1-based)

        Returns:
            Tuple of (min_row, min_col, max_row, max_col) or None if not merged
        """
        for merged_range in sheet.merged_cells.ranges:
            min_col, min_row, max_col, max_row = range_boundaries(str(merged_range))
            if min_row <= row <= max_row and min_col <= col <= max_col:
                return (min_row, min_col, max_row, max_col)
        return None

    def _is_cell_in_merge(self, sheet: Worksheet, row: int, col: int) -> bool:
        """Check if a cell is part of a merged range."""
        return self._get_merged_cell_range(sheet, row, col) is not None

    def _get_merged_cell_value(self, sheet: Worksheet, row: int, col: int) -> Any:
        """
        Get the value from a merged cell (reads from top-left cell of merge range).

        Args:
            sheet: Worksheet
            row: Row index (1-based)
            col: Column index (1-based)

        Returns:
            Cell value (may be None)
        """
        merge_range = self._get_merged_cell_range(sheet, row, col)
        if merge_range:
            # Get value from top-left cell of merged range
            min_row, min_col, _, _ = merge_range
            return sheet.cell(min_row, min_col).value
        else:
            # Not merged, return cell value directly
            return sheet.cell(row, col).value


    def _infer_cell_type(self, cell: Cell) -> str:
        """
        Infer the data type expected in a cell.

        Args:
            cell: openpyxl Cell object

        Returns:
            Type string: 'number', 'date', 'currency', 'percentage', 'text'
        """
        # Check number format
        if cell.number_format:
            fmt = cell.number_format.lower()

            if any(x in fmt for x in ['$', 'usd', 'currency']):
                return 'currency'
            elif '%' in fmt:
                return 'percentage'
            elif any(x in fmt for x in ['m/d/y', 'd/m/y', 'yyyy', 'date']):
                return 'date'
            elif any(x in fmt for x in ['0', '#', 'number']):
                return 'number'

        # Check cell value type
        if isinstance(cell.value, (int, float)):
            return 'number'

        # Default to text
        return 'text'

    def _get_cell_display_value(self, cell: Cell) -> str:
        """Get the display value of a cell as a string."""
        if cell.value is None:
            return ""
        return str(cell.value)

    def _detect_formula_cells(self, sheet: Worksheet) -> List[str]:
        """
        Detect all cells containing formulas.

        Args:
            sheet: openpyxl Worksheet object

        Returns:
            List of cell coordinates (e.g., ["D10", "E20"])
        """
        formula_cells = []

        for row in sheet.iter_rows():
            for cell in row:
                if cell.data_type == 'f':  # Formula cell
                    formula_cells.append(cell.coordinate)

        return formula_cells
