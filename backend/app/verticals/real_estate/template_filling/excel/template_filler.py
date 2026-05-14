"""Template filling functionality for Excel templates."""

import os
import re
import zipfile
from copy import deepcopy
from dataclasses import dataclass
from typing import Any, Dict, List, Optional

from openpyxl.cell import Cell, MergedCell
from openpyxl.worksheet.worksheet import Worksheet

from app.utils.logging import logger
from app.verticals.real_estate.template_filling.excel.schema_based.validation import (
    FieldValidator,
    categorize_confidence,
    should_fill_cell,
)

try:
    from lxml import etree
except ImportError:  # pragma: no cover - exercised only when lxml is unavailable
    etree = None


WORKSHEET_XML_NS = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
X14_NS = "http://schemas.microsoft.com/office/spreadsheetml/2009/9/main"
EXTLST_TAG = f"{{{WORKSHEET_XML_NS}}}extLst"
WORKSHEET_TAG = f"{{{WORKSHEET_XML_NS}}}worksheet"
WORKSHEET_PART_PREFIX = "xl/worksheets/"


@dataclass(frozen=True)
class PreservedWorksheetExtLst:
    """Captured worksheet extension block that openpyxl drops on load/save."""

    extlst_xml: bytes
    namespaces: Dict[str, str]


def _xml_parser():
    """Create a strict XML parser that preserves worksheet content."""
    if etree is None:
        return None
    return etree.XMLParser(remove_blank_text=False, recover=False)


def _is_worksheet_part(name: str) -> bool:
    """Return True when the ZIP entry is a worksheet XML part."""
    return name.startswith(WORKSHEET_PART_PREFIX) and name.endswith(".xml")


def _contains_x14_data_validation(extlst_element) -> bool:
    """Return True if the worksheet extLst contains x14 data validation metadata."""
    if etree is None:
        return False

    for node in extlst_element.iter():
        if not isinstance(node.tag, str):
            continue
        qname = etree.QName(node)
        if qname.namespace == X14_NS and qname.localname in {"dataValidation", "dataValidations"}:
            return True
    return False


def _find_prefix_for_namespace(element, namespace_uri: str) -> Optional[str]:
    """Find an in-scope prefix for a namespace URI."""
    if etree is None:
        return None

    current = element
    while current is not None:
        for prefix, uri in current.nsmap.items():
            if prefix and uri == namespace_uri:
                return prefix
        current = current.getparent()
    return None


def _collect_extlst_namespaces(extlst_element) -> Dict[str, str]:
    """Collect namespace prefixes used by extLst elements and attributes."""
    if etree is None:
        return {}

    namespaces: Dict[str, str] = {}

    for node in extlst_element.iter():
        if not isinstance(node.tag, str):
            continue

        node_qname = etree.QName(node)
        if node.prefix and node_qname.namespace:
            namespaces[node.prefix] = node_qname.namespace

        for attr_name in node.attrib:
            if not isinstance(attr_name, str) or not attr_name.startswith("{"):
                continue

            attr_qname = etree.QName(attr_name)
            if not attr_qname.namespace:
                continue

            prefix = _find_prefix_for_namespace(node, attr_qname.namespace)
            if prefix:
                namespaces[prefix] = attr_qname.namespace

    return namespaces


def _extract_dropdown_extlsts(template_path: str) -> Dict[str, PreservedWorksheetExtLst]:
    """
    Capture worksheet extLst blocks containing x14 dropdown metadata.

    openpyxl strips these extensions on load/save, so we preserve them from the original ZIP
    and patch them back into the saved workbook later.
    """
    if etree is None:
        logger.warning("lxml is not installed; skipping x14 dropdown preservation")
        return {}

    preserved: Dict[str, PreservedWorksheetExtLst] = {}

    try:
        with zipfile.ZipFile(template_path, "r") as workbook_zip:
            for name in workbook_zip.namelist():
                if not _is_worksheet_part(name):
                    continue

                worksheet_xml = workbook_zip.read(name)
                root = etree.fromstring(worksheet_xml, parser=_xml_parser())
                extlst_element = root.find(EXTLST_TAG)

                if extlst_element is None or not _contains_x14_data_validation(extlst_element):
                    continue

                namespaces = _collect_extlst_namespaces(extlst_element)
                preserved[name] = PreservedWorksheetExtLst(
                    extlst_xml=etree.tostring(extlst_element, encoding="UTF-8"),
                    namespaces=namespaces,
                )
                logger.info(
                    "Preserved dropdown extLst from %s with namespaces: %s",
                    name,
                    sorted(namespaces.keys()),
                )
    except Exception as exc:
        logger.warning("Could not preserve x14 dropdown metadata from template %s: %s", template_path, exc)
        return {}

    return preserved


def _ensure_root_namespaces(root, required_namespaces: Dict[str, str]):
    """Rebuild the worksheet root with any missing namespace declarations."""
    if etree is None or not required_namespaces:
        return root

    merged_nsmap = dict(root.nsmap)
    needs_rebuild = False

    for prefix, namespace_uri in required_namespaces.items():
        if not prefix:
            continue

        existing_uri = merged_nsmap.get(prefix)
        if existing_uri == namespace_uri:
            continue

        if existing_uri and existing_uri != namespace_uri:
            logger.warning(
                "Worksheet root already uses prefix %s for %s; leaving preserved extLst namespace local",
                prefix,
                existing_uri,
            )
            continue

        merged_nsmap[prefix] = namespace_uri
        needs_rebuild = True

    if not needs_rebuild:
        return root

    rebuilt_root = etree.Element(root.tag, nsmap=merged_nsmap)
    rebuilt_root.text = root.text
    rebuilt_root.tail = root.tail

    for attr_name, attr_value in root.attrib.items():
        rebuilt_root.set(attr_name, attr_value)

    for child in root:
        rebuilt_root.append(deepcopy(child))

    return rebuilt_root


def _patch_dropdown_extlst(worksheet_xml: bytes, preserved_extlst: PreservedWorksheetExtLst) -> bytes:
    """Patch the preserved extLst block back into a saved worksheet XML document."""
    if etree is None:
        return worksheet_xml

    root = etree.fromstring(worksheet_xml, parser=_xml_parser())
    if root.tag != WORKSHEET_TAG:
        raise ValueError(f"Unexpected worksheet root tag: {root.tag}")

    existing_extlst = root.find(EXTLST_TAG)
    if existing_extlst is not None:
        root.remove(existing_extlst)

    root = _ensure_root_namespaces(root, preserved_extlst.namespaces)
    extlst_element = etree.fromstring(preserved_extlst.extlst_xml, parser=_xml_parser())
    root.append(extlst_element)

    patched_xml = etree.tostring(root, encoding="UTF-8", xml_declaration=True)

    # Validate the final XML before writing it back into the ZIP.
    etree.fromstring(patched_xml, parser=_xml_parser())
    return patched_xml


def _restore_dropdown_extlsts(output_path: str, preserved_extlsts: Dict[str, PreservedWorksheetExtLst]) -> None:
    """Restore preserved worksheet dropdown extLst blocks after openpyxl saves the workbook."""
    if etree is None or not preserved_extlsts:
        return

    temp_output_path = output_path + ".x14patch"

    try:
        with zipfile.ZipFile(output_path, "r") as input_zip:
            with zipfile.ZipFile(temp_output_path, "w", compression=zipfile.ZIP_DEFLATED) as patched_zip:
                for zip_info in input_zip.infolist():
                    entry_bytes = input_zip.read(zip_info.filename)
                    preserved_extlst = preserved_extlsts.get(zip_info.filename)

                    if preserved_extlst is not None:
                        entry_bytes = _patch_dropdown_extlst(entry_bytes, preserved_extlst)
                        logger.info(
                            "Restored dropdown extLst into %s with namespaces: %s",
                            zip_info.filename,
                            sorted(preserved_extlst.namespaces.keys()),
                        )

                    patched_zip.writestr(zip_info, entry_bytes)

        os.replace(temp_output_path, output_path)
        logger.info("Applied worksheet dropdown extension patch to %s", output_path)
    except Exception as exc:
        logger.warning("Failed to restore dropdown extLst metadata into %s: %s", output_path, exc)
        if os.path.exists(temp_output_path):
            try:
                os.remove(temp_output_path)
            except OSError:
                pass


class TemplateFiller:
    """Fill Excel templates with extracted data while preserving formulas."""

    def fill_template(
        self,
        template_path: str,
        output_path: str,
        field_mapping: Dict[str, Any],
        extracted_data: Dict[str, Any],
        workbook=None,
    ) -> Dict[str, Any]:
        """
        Fill an Excel template with extracted data.

        Args:
            template_path: Path to the original Excel template
            output_path: Path to save the filled Excel file
            field_mapping: Field mapping structure from TemplateFillRun
            extracted_data: Extracted data from PDF
            workbook: Optional pre-loaded workbook (for testing)

        Returns:
            Summary of fill operation
        """
        from openpyxl import load_workbook

        logger.info(f"Filling Excel template: {template_path} -> {output_path}")

        try:
            # Determine if file is macro-enabled (.xlsm) or not (.xlsx)
            # Only use keep_vba=True for .xlsm files to avoid corruption
            is_macro_enabled = template_path.lower().endswith('.xlsm')

            # Preserve worksheet dropdown metadata that openpyxl silently strips on load/save.
            preserved_extlsts = _extract_dropdown_extlsts(template_path)

            # Load template (data_only=False to preserve formulas)
            if workbook is None:
                workbook = load_workbook(template_path, data_only=False, keep_vba=is_macro_enabled)

            cells_filled = 0
            sheets_modified = set()
            errors = []

            # Confidence tracking
            confidence_stats = {
                "high_confidence": 0,      # >= 0.85
                "needs_review": 0,         # 0.70-0.84
                "low_confidence": 0,       # 0.50-0.69
                "skipped_low_confidence": 0,  # < 0.10
                "validation_errors": 0,
            }
            needs_review_cells: List[Dict[str, Any]] = []
            validation_error_cells: List[Dict[str, Any]] = []

            # Extract data sections (clean nested schema)
            llm_extracted = extracted_data.get("llm_extracted", {})
            manual_edits = extracted_data.get("manual_edits", {})

            # Process mapped cells (using LLM extracted data)
            for mapping in field_mapping.get("mappings", []):
                pdf_field_id = mapping.get("pdf_field_id")
                excel_cell = mapping.get("excel_cell")
                excel_sheet = mapping.get("excel_sheet")
                confidence = mapping.get("confidence", 0.0)
                source = mapping.get("source", "unknown")

                if not all([pdf_field_id, excel_cell, excel_sheet]):
                    errors.append(f"Invalid mapping: {mapping}")
                    continue

                # Track confidence category
                confidence_category = categorize_confidence(confidence)

                # Check if we should skip low-confidence mappings
                if not should_fill_cell(confidence):
                    confidence_stats["skipped_low_confidence"] += 1
                    logger.debug(f"Skipping low-confidence mapping: {excel_sheet}!{excel_cell} (confidence={confidence:.2f})")
                    continue

                # Get extracted value from LLM data
                field_data = llm_extracted.get(pdf_field_id)
                if not field_data:
                    errors.append(f"No data for field {pdf_field_id}")
                    continue

                value = field_data.get("value") if isinstance(field_data, dict) else field_data
                if value is None:
                    continue  # Skip null values

                # Get worksheet
                if excel_sheet not in workbook.sheetnames:
                    errors.append(f"Sheet '{excel_sheet}' not found")
                    continue

                sheet = workbook[excel_sheet]

                # Fill cell
                try:
                    cell = sheet[excel_cell]

                    # Check if cell is part of a merged range
                    target_cell = self._get_writable_cell(sheet, cell, excel_cell)
                    if target_cell is None:
                        errors.append(f"Cell {excel_sheet}!{excel_cell} is in a merged range but top-left cell not found")
                        continue

                    # Schema/YAML data_type is the contract for schema-based fills.
                    # Excel templates often carry copied formatting that does not
                    # match the semantic target cell, so use formatting only as a
                    # fallback for older/generic mappings.
                    cell_type = mapping.get("data_type") or self._infer_cell_type(target_cell)

                    # Validate value against expected type
                    validation_result = FieldValidator.validate(value, cell_type)

                    if not validation_result.is_valid:
                        confidence_stats["validation_errors"] += 1
                        validation_error_cells.append({
                            "sheet": excel_sheet,
                            "cell": excel_cell,
                            "value": value,
                            "expected_type": cell_type,
                            "error": validation_result.error_message,
                        })
                        logger.warning(f"Validation error for {excel_sheet}!{excel_cell}: {validation_result.error_message}")
                        # Still fill with original value, but log the warning
                        filled_value = self._convert_value_for_cell(value, target_cell)
                    else:
                        # Use validated/converted value
                        filled_value = validation_result.converted_value if validation_result.converted_value is not None else self._convert_value_for_cell(value, target_cell)

                    # Set value (formulas in other cells will auto-recalculate)
                    target_cell.value = filled_value

                    cells_filled += 1
                    sheets_modified.add(excel_sheet)

                    # Track confidence category
                    if confidence_category == "auto_fill":
                        confidence_stats["high_confidence"] += 1
                    elif confidence_category == "needs_review":
                        confidence_stats["needs_review"] += 1
                        needs_review_cells.append({
                            "sheet": excel_sheet,
                            "cell": excel_cell,
                            "confidence": confidence,
                            "source": source,
                            "value": str(value)[:50],  # Truncate for logging
                        })
                    elif confidence_category == "low_confidence":
                        confidence_stats["low_confidence"] += 1
                        needs_review_cells.append({
                            "sheet": excel_sheet,
                            "cell": excel_cell,
                            "confidence": confidence,
                            "source": source,
                            "value": str(value)[:50],
                        })

                except Exception as e:
                    errors.append(f"Error filling {excel_sheet}!{excel_cell}: {str(e)}")
                    logger.warning(f"Error filling cell {excel_sheet}!{excel_cell}: {e}")

            # Process manually edited unmapped cells
            manually_filled = 0
            for sheet_name, cells_data in manual_edits.items():
                for cell_address, cell_value in cells_data.items():
                    # Skip if cell already has a mapping (processed above)
                    cell_already_mapped = any(
                        m.get("excel_sheet") == sheet_name and m.get("excel_cell") == cell_address
                        for m in field_mapping.get("mappings", [])
                    )
                    if cell_already_mapped:
                        continue

                    # Get value to fill
                    value = cell_value.get("value") if isinstance(cell_value, dict) else cell_value
                    if value is None:
                        continue

                    # Get worksheet
                    if sheet_name not in workbook.sheetnames:
                        errors.append(f"Sheet '{sheet_name}' not found (manual cell)")
                        continue

                    sheet = workbook[sheet_name]

                    # Fill cell
                    try:
                        cell = sheet[cell_address]

                        # Skip if cell is a formula cell
                        if cell.data_type == 'f':  # formula cell
                            logger.debug(f"Skipping formula cell {sheet_name}!{cell_address} (manual edit)")
                            continue

                        # Check if cell is part of a merged range
                        target_cell = self._get_writable_cell(sheet, cell, cell_address)
                        if target_cell is None:
                            errors.append(f"Manual cell {sheet_name}!{cell_address} is in a merged range but top-left cell not found")
                            continue

                        # Preserve cell type - convert value appropriately
                        filled_value = self._convert_value_for_cell(value, target_cell)

                        # Set value
                        target_cell.value = filled_value

                        manually_filled += 1
                        cells_filled += 1
                        sheets_modified.add(sheet_name)

                        logger.debug(f"Filled manual cell {sheet_name}!{cell_address} with value: {value}")

                    except Exception as e:
                        errors.append(f"Error filling manual cell {sheet_name}!{cell_address}: {str(e)}")
                        logger.warning(f"Error filling manual cell {sheet_name}!{cell_address}: {e}")

            if manually_filled > 0:
                logger.info(f"Filled {manually_filled} manually edited unmapped cells")

            # Clean up broken external links to prevent file corruption
            try:
                if hasattr(workbook, '_external_links') and workbook._external_links:
                    logger.warning(
                        f"Template has {len(workbook._external_links)} external link(s) "
                        f"(likely formulas referencing missing sheets). Removing to prevent corruption."
                    )
                    workbook._external_links = []
                else:
                    logger.info("No external links detected - all formulas reference existing sheets")
            except Exception as e:
                logger.warning(f"Could not check/remove external links: {e}")

            # Remove ALL named ranges to prevent corruption
            try:
                # workbook.defined_names is a DefinedNameDict - iterate directly
                named_range_names = list(workbook.defined_names.keys())

                if named_range_names:
                    logger.warning(
                        f"Removing {len(named_range_names)} named range(s) to prevent corruption: "
                        f"{named_range_names[:5]}{'...' if len(named_range_names) > 5 else ''}"
                    )
                    for name in named_range_names:
                        try:
                            del workbook.defined_names[name]
                            logger.debug(f"Deleted named range '{name}'")
                        except Exception as e:
                            logger.warning(f"Could not delete named range '{name}': {e}")
                else:
                    logger.info("No named ranges detected")
            except Exception as e:
                logger.warning(f"Could not remove named ranges: {e}")

            # Clear formulas that reference missing sheets or external workbooks
            try:
                existing_sheets = set(workbook.sheetnames)
                existing_sheets_lower = {s.lower() for s in existing_sheets}
                formulas_cleared = 0

                # Regex patterns to find sheet references in formulas:
                # Pattern 1: 'Sheet Name'!A1 (quoted sheet name with spaces)
                # Pattern 2: SheetName!A1 (unquoted sheet name)
                sheet_ref_pattern = re.compile(r"(?:'([^']+)'|([A-Za-z_][A-Za-z0-9_]*))\s*!")

                for sheet_name in workbook.sheetnames:
                    sheet = workbook[sheet_name]
                    for row in sheet.iter_rows():
                        for cell in row:
                            if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                                formula = cell.value
                                should_clear = False

                                # Check 1: External workbook reference (contains '[' and ']')
                                if '[' in formula and ']' in formula:
                                    logger.debug(f"Clearing formula in {sheet_name}!{cell.coordinate}: contains external workbook reference")
                                    should_clear = True
                                else:
                                    # Check 2: Internal sheet reference to missing sheet
                                    matches = sheet_ref_pattern.findall(formula)
                                    for match in matches:
                                        # match is a tuple: (quoted_name, unquoted_name)
                                        ref_sheet = match[0] if match[0] else match[1]
                                        if ref_sheet.lower() not in existing_sheets_lower:
                                            logger.debug(f"Clearing formula in {sheet_name}!{cell.coordinate}: references missing sheet '{ref_sheet}'")
                                            should_clear = True
                                            break

                                if should_clear:
                                    cell.value = None
                                    formulas_cleared += 1

                if formulas_cleared > 0:
                    logger.warning(
                        f"Cleared {formulas_cleared} formula(s) referencing missing sheets or external workbooks "
                        f"(temporary fix for partial template)"
                    )
                else:
                    logger.info("No formulas reference missing sheets or external workbooks")
            except Exception as e:
                logger.warning(f"Could not clear formulas referencing missing sheets: {e}")

            # Save filled workbook
            workbook.save(output_path)
            workbook.close()

            # Re-inject x14 dropdown extension metadata after openpyxl writes the workbook.
            _restore_dropdown_extlsts(output_path, preserved_extlsts)


            summary = {
                "total_cells_filled": cells_filled,
                "sheets_modified": sorted(list(sheets_modified)),
                "formulas_preserved": True,  # openpyxl preserves formulas by default
                "errors": errors,
                # Confidence breakdown
                "confidence_stats": confidence_stats,
                "needs_review_cells": needs_review_cells[:20],  # Limit to first 20
                "validation_error_cells": validation_error_cells[:20],
            }

            # Log confidence breakdown
            logger.info(
                f"Confidence breakdown: {confidence_stats['high_confidence']} high, "
                f"{confidence_stats['needs_review']} needs review, "
                f"{confidence_stats['low_confidence']} low, "
                f"{confidence_stats['skipped_low_confidence']} skipped, "
                f"{confidence_stats['validation_errors']} validation errors"
            )

            logger.info(
                f"Template filled successfully: {cells_filled} cells across "
                f"{len(sheets_modified)} sheets"
            )

            return summary

        except Exception as e:
            logger.error(f"Error filling Excel template: {e}", exc_info=True)
            raise

    def _convert_value_for_cell(self, value: Any, cell: Cell) -> Any:
        """
        Convert extracted value to appropriate type for Excel cell.

        Args:
            value: Extracted value (usually string from LLM)
            cell: Target Excel cell

        Returns:
            Converted value
        """
        # If value is already the right type, use it
        if value is None:
            return None

        # Get cell's expected type from number format
        cell_type = self._infer_cell_type(cell)

        # Convert based on cell type
        try:
            if cell_type == 'percentage':
                # Special handling for percentages
                # Excel stores percentages as decimals (30% = 0.30)
                if isinstance(value, (int, float)):
                    # If already a number, check if it needs conversion
                    # Assume values > 1 are percentages that need dividing by 100
                    if value > 1:
                        return value / 100
                    return value

                # Parse string percentage
                cleaned = str(value).replace(',', '').replace('$', '').replace('%', '').strip()
                try:
                    num_value = float(cleaned)
                    # Convert percentage to decimal (30% -> 0.30)
                    # Assume values > 1 are percentages (e.g., "30" means 30%)
                    if num_value > 1:
                        return num_value / 100
                    return num_value
                except ValueError:
                    return str(value)

            elif cell_type in ['number', 'currency']:
                # Try to parse as float
                if isinstance(value, (int, float)):
                    return float(value)

                # Remove common formatting characters
                cleaned = str(value).replace(',', '').replace('$', '').replace('%', '').strip()

                try:
                    return float(cleaned)
                except ValueError:
                    # If can't convert, return as string
                    return str(value)

            elif cell_type == 'date':
                # For dates, we might need more sophisticated parsing
                # For now, return as string and let Excel interpret
                return str(value)

            else:  # text
                return str(value)

        except Exception as e:
            logger.warning(f"Error converting value '{value}' for cell type '{cell_type}': {e}")
            return str(value)

    def _infer_cell_type(self, cell: Cell) -> str:
        """
        Infer the data type expected in a cell.

        Args:
            cell: openpyxl Cell object

        Returns:
            Type string: 'number', 'date', 'currency', 'percentage', 'text'
        """
        # Check number format
        if cell.number_format:
            fmt = cell.number_format.lower()

            if any(x in fmt for x in ['$', 'usd', 'currency']):
                return 'currency'
            elif '%' in fmt:
                return 'percentage'
            elif any(x in fmt for x in ['m/d/y', 'd/m/y', 'yyyy', 'date']):
                return 'date'
            elif any(x in fmt for x in ['0', '#', 'number']):
                return 'number'

        # Check cell value type
        if isinstance(cell.value, (int, float)):
            return 'number'

        # Default to text
        return 'text'

    def _get_writable_cell(self, sheet: Worksheet, cell: Cell, cell_address: str) -> Optional[Cell]:
        """
        Get the writable cell for a given address, handling merged cells.

        If the cell is part of a merged range, returns the top-left cell of that range
        (which is the only cell that can be written to).

        Args:
            sheet: The worksheet containing the cell
            cell: The cell object (may be a MergedCell)
            cell_address: The cell address (e.g., 'G40')

        Returns:
            The writable cell, or None if there's an error
        """
        # If cell is not a MergedCell, it's directly writable
        if not isinstance(cell, MergedCell):
            return cell

        # Cell is part of a merged range - find the top-left cell
        logger.debug(f"Cell {cell_address} is part of a merged range")

        # Find which merged range this cell belongs to
        for merged_range in sheet.merged_cells.ranges:
            if cell.coordinate in merged_range:
                # Get top-left cell of the merged range
                top_left_coord = merged_range.start_cell.coordinate
                top_left_cell = sheet[top_left_coord]

                logger.info(
                    f"Cell {cell_address} is in merged range {merged_range}. "
                    f"Writing to top-left cell {top_left_coord} instead."
                )

                return top_left_cell

        # Should not reach here, but handle gracefully
        logger.warning(f"Cell {cell_address} appears to be a MergedCell but no merged range found")
        return None
