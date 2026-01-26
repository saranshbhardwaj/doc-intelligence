"""Table detection using Excel native, border, and heuristic methods."""

from typing import Any, Dict, List, Optional, Tuple
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import range_boundaries
from openpyxl.worksheet.worksheet import Worksheet

from app.utils.logging import logger
from .style_inspector import StyleInspector


class TableDetector:
    """Detect table structures in Excel worksheets using 3-layer approach."""

    def __init__(self):
        """Initialize table detector with style inspector."""
        self._style_inspector = StyleInspector()

    def _regions_overlap(self, region1: Dict[str, int], region2: Dict[str, int]) -> bool:
        """
        Check if two table regions overlap.

        Args:
            region1: Dict with start_row, end_row, start_col, end_col
            region2: Dict with start_row, end_row, start_col, end_col

        Returns:
            True if regions share any cells
        """
        # Check if rows overlap
        rows_overlap = not (
            region1['end_row'] < region2['start_row'] or
            region2['end_row'] < region1['start_row']
        )

        # Check if columns overlap
        cols_overlap = not (
            region1['end_col'] < region2['start_col'] or
            region2['end_col'] < region1['start_col']
        )

        return rows_overlap and cols_overlap

    def _detect_excel_native_tables(self, sheet: Worksheet) -> List[Dict[str, Any]]:
        """
        Layer 1: Detect Excel native tables (Insert → Table).

        These are explicitly defined tables that Excel stores in the worksheet.

        Args:
            sheet: Worksheet

        Returns:
            List of table metadata for native Excel tables
        """
        tables = []

        if not hasattr(sheet, 'tables') or not sheet.tables:
            return tables

        for table_name, table_obj in sheet.tables.items():
            # Parse table range (e.g., "A1:E10")
            min_col, min_row, max_col, max_row = range_boundaries(table_obj.ref)

            logger.info(
                f"📋 Layer 1: Found Excel native table '{table_name}' in '{sheet.title}': "
                f"rows {min_row}-{max_row}, columns {min_col}-{max_col}"
            )

            # Extract header row (first row of the table)
            header_row = min_row
            header_cells = []

            for col_idx in range(min_col, max_col + 1):
                cell = sheet.cell(header_row, col_idx)
                if cell.value:
                    header_value = str(cell.value).strip()
                    header_cells.append((col_idx, header_value))

            # Analyze table structure
            if header_cells:
                table = self._analyze_table_structure(sheet, header_row, header_cells, end_row=max_row)
                if table:
                    table['table_name'] = table_name  # Use native table name
                    table['detection_method'] = 'native'
                    tables.append(table)

        return tables

    def _detect_bordered_tables(
        self,
        sheet: Worksheet,
        excluded_regions: List[Dict[str, int]]
    ) -> List[Dict[str, Any]]:
        """
        Layer 2: Detect tables by analyzing cell borders.

        Looks for contiguous regions with borders that form table structures.

        Args:
            sheet: Worksheet
            excluded_regions: Regions already detected by higher-priority layers

        Returns:
            List of table metadata for bordered tables
        """
        tables = []
        max_scan_row = min(sheet.max_row, 100)
        max_scan_col = min(sheet.max_column, 50)

        visited_rows = set()

        for row_idx in range(1, max_scan_row + 1):
            if row_idx in visited_rows:
                continue

            # Look for rows with borders that might be table headers
            row_cells = []
            has_borders = False

            for col_idx in range(1, max_scan_col + 1):
                cell = sheet.cell(row_idx, col_idx)
                if cell.value is not None and cell.data_type != 'f':
                    header_value = str(cell.value).strip()
                    if len(header_value) < 100:
                        row_cells.append((col_idx, header_value))

                if self._style_inspector._has_border(cell, 'any'):
                    has_borders = True

            # Need at least 3 cells and some borders to consider as table
            if len(row_cells) >= 3 and has_borders:
                start_col = min(c[0] for c in row_cells)
                end_col = max(c[0] for c in row_cells)

                # Verify this looks like a header using scoring (optional quality check)
                header_score = self._calculate_header_score(sheet, row_idx, start_col, end_col)

                # For border-based detection, use lower threshold since borders are already strong signal
                if header_score < 0.15:
                    # Skip rows with very low header probability (likely false positive)
                    logger.debug(f"Skipping bordered row {row_idx}: low header score {header_score:.2f}")
                    continue

                # Find extent of bordered region below this row
                end_row = row_idx
                for check_row in range(row_idx + 1, min(row_idx + 51, max_scan_row + 1)):
                    has_content_or_border = False
                    for col_idx in range(start_col, end_col + 1):
                        cell = sheet.cell(check_row, col_idx)
                        if cell.value is not None or self._style_inspector._has_border(cell, 'any'):
                            has_content_or_border = True
                            break

                    if has_content_or_border:
                        end_row = check_row
                    else:
                        break

                # Check if this region overlaps with any excluded region
                region = {
                    'start_row': row_idx,
                    'end_row': end_row,
                    'start_col': start_col,
                    'end_col': end_col,
                }

                is_excluded = any(
                    self._regions_overlap(region, excluded)
                    for excluded in excluded_regions
                )

                if not is_excluded and end_row > row_idx:
                    # Analyze as table
                    table = self._analyze_table_structure(sheet, row_idx, row_cells, end_row=end_row)
                    if table:
                        table['detection_method'] = 'border'
                        tables.append(table)
                        logger.info(
                            f"🔲 Layer 2: Found bordered table in '{sheet.title}': "
                            f"rows {row_idx}-{end_row}, columns {start_col}-{end_col}"
                        )

                    # Mark rows as visited
                    for r in range(row_idx, end_row + 1):
                        visited_rows.add(r)

        return tables

    def _split_row_cells_by_gaps(self, row_cells: List[tuple], gap_threshold: int = 1) -> List[List[tuple]]:
        """
        Split row cells into separate groups when there are gaps (empty columns).

        This prevents combining separate sections (e.g., key-value section + table)
        into a single table, enabling proper hierarchical header detection.

        Args:
            row_cells: List of (col_idx, value) tuples
            gap_threshold: Minimum gap size to split (default 1 empty column)

        Returns:
            List of row cell groups (each group is a list of (col_idx, value) tuples)
        """
        if not row_cells:
            return []

        # Sort by column index
        sorted_cells = sorted(row_cells, key=lambda x: x[0])

        groups = []
        current_group = [sorted_cells[0]]

        for i in range(1, len(sorted_cells)):
            prev_col = sorted_cells[i - 1][0]
            curr_col = sorted_cells[i][0]
            gap = curr_col - prev_col - 1  # Number of empty columns between

            if gap >= gap_threshold:
                # Gap detected - save current group and start new one
                groups.append(current_group)
                current_group = [sorted_cells[i]]
            else:
                # No significant gap - add to current group
                current_group.append(sorted_cells[i])

        # Add the last group
        if current_group:
            groups.append(current_group)

        return groups

    def _calculate_header_score(
        self,
        sheet: Worksheet,
        row_idx: int,
        start_col: int,
        end_col: int
    ) -> float:
        """
        Calculate probability that a row is a table header using combined signals.

        Returns score 0.0-1.0 based on multiple styling and structural signals.

        Scoring weights:
        - Bold text: 0.30 (strong signal)
        - Background fill: 0.30 (strong signal)
        - Bottom border: 0.20 (common header pattern)
        - Consistent styling across row: 0.15 (indicates intentional header)
        - Text content: 0.10 (has non-numeric values)
        - Frozen row: 0.25 (structural signal)
        - Data validation in columns below: 0.20 (indicates structured input)

        Args:
            sheet: Worksheet
            row_idx: Row to evaluate
            start_col: Starting column
            end_col: Ending column

        Returns:
            Score from 0.0 (not a header) to 1.0 (definitely a header)
        """
        if end_col < start_col:
            return 0.0

        score = 0.0
        cell_count = 0
        text_count = 0
        bold_count = 0
        fill_count = 0
        bottom_border_count = 0
        validation_count = 0

        for col_idx in range(start_col, end_col + 1):
            cell = sheet.cell(row_idx, col_idx)
            cell_count += 1

            # Check styling signals
            if self._style_inspector._is_bold(cell):
                bold_count += 1

            if self._style_inspector._has_fill(cell):
                fill_count += 1

            if self._style_inspector._has_border(cell, 'bottom'):
                bottom_border_count += 1

            # Check if has text content (not just numbers)
            if cell.value:
                val_str = str(cell.value).strip()
                if val_str and not val_str.replace('.', '').replace('-', '').isdigit():
                    text_count += 1

            # Check for data validation in next few rows (indicates header)
            has_validation_below = False
            for check_row in range(row_idx + 1, min(row_idx + 6, sheet.max_row + 1)):
                if self._style_inspector._has_data_validation(sheet, check_row, col_idx):
                    has_validation_below = True
                    break
            if has_validation_below:
                validation_count += 1

        if cell_count == 0:
            return 0.0

        # Calculate component scores
        bold_score = (bold_count / cell_count) * 0.30
        fill_score = (fill_count / cell_count) * 0.30
        border_score = (bottom_border_count / cell_count) * 0.20
        text_score = (text_count / cell_count) * 0.10
        validation_score = (validation_count / cell_count) * 0.20

        # Check for consistent styling (bonus)
        consistent_styling_bonus = 0.0
        if self._style_inspector._row_has_consistent_styling(sheet, row_idx, start_col, end_col):
            consistent_styling_bonus = 0.15

        # Check if row is frozen (strong signal)
        frozen_rows = self._style_inspector._get_frozen_rows(sheet)
        frozen_bonus = 0.0
        if frozen_rows > 0 and row_idx <= frozen_rows:
            frozen_bonus = 0.25

        # Combine scores
        score = bold_score + fill_score + border_score + text_score + validation_score + consistent_styling_bonus + frozen_bonus

        # Cap at 1.0
        return min(score, 1.0)

    def _detect_tables_heuristic(
        self,
        sheet: Worksheet,
        excluded_regions: List[Dict[str, int]]
    ) -> List[Dict[str, Any]]:
        """
        Layer 3: Detect tables using heuristics (fallback method).

        Enhanced support for complex tables:
        1. Simple tables: headers in one row, data below
        2. Multi-level hierarchical headers: detects up to 1 parent header row above main headers (max 2 levels)
        3. Merged cells: properly handles merged header cells (e.g., "Current Rent" spanning 3 columns)
        4. Hierarchical column names: builds names like "Current Rent | Average" for disambiguation
        5. Gap detection: splits sections separated by empty columns into separate tables
        6. Smart header scoring: uses font styling (bold, fill), borders, frozen panes, data validation

        Args:
            sheet: openpyxl Worksheet object
            excluded_regions: Regions already detected by higher-priority layers

        Returns:
            List of table metadata
        """
        tables = []

        # Scan for potential header rows (rows with multiple consecutive text cells)
        max_scan_row = min(sheet.max_row, 100)  # Focus on first 100 rows
        max_scan_col = min(sheet.max_column, 50)

        for row_idx in range(1, max_scan_row + 1):
            # Check if this row looks like headers
            row_cells = []
            for col_idx in range(1, max_scan_col + 1):
                cell = sheet.cell(row_idx, col_idx)
                # Accept strings, numbers, dates as potential headers (but not formulas)
                if cell.value is not None and cell.data_type != 'f':
                    # Convert to string for consistent handling
                    header_value = str(cell.value).strip()
                    # Skip very long values (likely not headers)
                    if len(header_value) < 100:
                        row_cells.append((col_idx, header_value))

            # Need at least 3 cells to be a header row
            if len(row_cells) >= 3:
                # Split into separate sections if there are gaps (empty columns)
                # This prevents combining key-value sections with actual tables
                cell_groups = self._split_row_cells_by_gaps(row_cells, gap_threshold=1)

                # Analyze each group as a separate potential table
                for group in cell_groups:
                    if len(group) >= 3:  # Need at least 3 cells for a table
                        # Check if cells are somewhat consecutive within this group
                        col_indices = [c[0] for c in group]
                        if max(col_indices) - min(col_indices) <= len(col_indices) + 5:  # Allow some gaps
                            # Calculate header score using combined signals
                            start_col = min(col_indices)
                            end_col = max(col_indices)
                            header_score = self._calculate_header_score(sheet, row_idx, start_col, end_col)

                            # Use score-based threshold (0.3 = moderate confidence)
                            # Lower threshold than before to catch more tables, but with better quality signal
                            is_likely_header = header_score >= 0.3

                            # Log score for debugging (only if score is significant)
                            if header_score > 0.2:
                                logger.debug(
                                    f"Row {row_idx} cols {start_col}-{end_col}: header_score={header_score:.2f} "
                                    f"({'HEADER' if is_likely_header else 'skip'})"
                                )

                            if is_likely_header:
                                # Check if this overlaps with excluded regions
                                start_col = min(c[0] for c in group)
                                end_col = max(c[0] for c in group)

                                # Estimate table end row (scan for data below)
                                end_row = row_idx
                                for check_row in range(row_idx + 1, min(row_idx + 51, max_scan_row + 1)):
                                    has_content = False
                                    for col_idx in range(start_col, end_col + 1):
                                        cell = sheet.cell(check_row, col_idx)
                                        if cell.value is not None or self._style_inspector._has_border(cell, 'any') or self._style_inspector._has_fill(cell):
                                            has_content = True
                                            break
                                    if has_content:
                                        end_row = check_row
                                    else:
                                        break

                                region = {
                                    'start_row': row_idx,
                                    'end_row': end_row,
                                    'start_col': start_col,
                                    'end_col': end_col,
                                }

                                # Skip if overlaps with higher-priority detection
                                is_excluded = any(
                                    self._regions_overlap(region, excluded)
                                    for excluded in excluded_regions
                                )

                                if not is_excluded:
                                    # Potential table found!
                                    table = self._analyze_table_structure(sheet, row_idx, group, end_row=end_row)
                                    if table:
                                        table['detection_method'] = 'heuristic'
                                        tables.append(table)
                                        logger.info(
                                            f"🔍 Layer 3: Found heuristic table in '{sheet.title}': "
                                            f"rows {row_idx}-{end_row}, columns {start_col}-{end_col}"
                                        )

        # Remove duplicate/overlapping tables
        tables = self._deduplicate_tables(tables)

        return tables

    def detect_tables(self, sheet: Worksheet) -> List[Dict[str, Any]]:
        """
        Detect table structures using layered approach (priority-based).

        Layer 1 (Highest): Excel native tables (Insert → Table)
        Layer 2 (Medium): Border-based detection
        Layer 3 (Lowest): Heuristic detection (fallback)

        Args:
            sheet: openpyxl Worksheet object

        Returns:
            List of table metadata
        """
        all_tables = []
        excluded_regions = []

        # Layer 1: Excel native tables (highest priority)
        native_tables = self._detect_excel_native_tables(sheet)
        all_tables.extend(native_tables)

        # Add native table regions to excluded list
        for table in native_tables:
            excluded_regions.append({
                'start_row': table['start_row'],
                'end_row': table.get('end_row') or (table['start_row'] + len(table.get('data_rows', []))),
                'start_col': table['start_col'],
                'end_col': table['end_col'],
            })

        # Layer 2: Border-based detection (medium priority)
        bordered_tables = self._detect_bordered_tables(sheet, excluded_regions)
        all_tables.extend(bordered_tables)

        # Add bordered table regions to excluded list
        for table in bordered_tables:
            excluded_regions.append({
                'start_row': table['start_row'],
                'end_row': table.get('end_row') or (table['start_row'] + len(table.get('data_rows', []))),
                'start_col': table['start_col'],
                'end_col': table['end_col'],
            })

        # Layer 3: Heuristic detection (lowest priority, fallback)
        heuristic_tables = self._detect_tables_heuristic(sheet, excluded_regions)
        all_tables.extend(heuristic_tables)

        logger.info(
            f"Detected {len(all_tables)} tables in sheet '{sheet.title}' "
            f"(Layer1: {len(native_tables)}, Layer2: {len(bordered_tables)}, Layer3: {len(heuristic_tables)})"
        )

        return all_tables

    def _detect_header_rows(
        self,
        sheet: Worksheet,
        potential_header_row: int,
        start_col: int,
        end_col: int,
        get_merged_cell_value_fn
    ) -> List[int]:
        """
        Detect multiple header rows above a potential header row.

        Looks upward from the detected header row to find parent header rows
        (typically merged cells spanning multiple columns).

        Args:
            sheet: Worksheet
            potential_header_row: The row index that looks like the main header
            start_col: Starting column of table
            end_col: Ending column of table
            get_merged_cell_value_fn: Function to get merged cell values

        Returns:
            List of header row indices in order (top to bottom)
        """
        header_rows = [potential_header_row]

        # Look up to 1 row above for a parent header level (max 2 total levels)
        for offset in range(1, 2):
            check_row = potential_header_row - offset
            if check_row < 1:
                break

            # Check if this row has values that span multiple columns (merged cells)
            has_parent_headers = False
            for col_idx in range(start_col, end_col + 1):
                cell_value = get_merged_cell_value_fn(sheet, check_row, col_idx)
                if cell_value and str(cell_value).strip():
                    # Check if this cell is part of a merged range spanning multiple columns
                    merge_range = self._get_merged_cell_range_fn(sheet, check_row, col_idx)
                    if merge_range:
                        _, min_col, _, max_col = merge_range
                        # If merged range spans at least 2 columns, it's likely a parent header
                        if max_col - min_col >= 1:
                            has_parent_headers = True
                            break

            if has_parent_headers:
                header_rows.insert(0, check_row)  # Add to front (top-down order)
            else:
                # Stop looking if we hit a row without parent headers
                break

        return header_rows

    def _build_hierarchical_headers(
        self,
        sheet: Worksheet,
        header_rows: List[int],
        start_col: int,
        end_col: int,
        get_merged_cell_value_fn
    ) -> List[Dict[str, Any]]:
        """
        Build hierarchical column headers from multiple header rows.

        Args:
            sheet: Worksheet
            header_rows: List of header row indices (top to bottom)
            start_col: Starting column
            end_col: Ending column
            get_merged_cell_value_fn: Function to get merged cell values

        Returns:
            List of column header dictionaries with hierarchical names
        """
        column_headers = []

        for col_idx in range(start_col, end_col + 1):
            # Collect header values from all levels for this column
            header_parts = []

            for header_row_idx in header_rows:
                cell_value = get_merged_cell_value_fn(sheet, header_row_idx, col_idx)
                if cell_value:
                    header_text = str(cell_value).strip()
                    # Only add if not empty and not a duplicate of the last part
                    if header_text and (not header_parts or header_parts[-1] != header_text):
                        header_parts.append(header_text)

            # Build hierarchical name
            if header_parts:
                # Join with " | " to show hierarchy
                hierarchical_name = " | ".join(header_parts)
                # Also keep the leaf name (last part)
                leaf_name = header_parts[-1] if header_parts else ""
            else:
                hierarchical_name = ""
                leaf_name = ""

            column_headers.append({
                "col": col_idx,
                "col_letter": get_column_letter(col_idx),
                "header": hierarchical_name,  # Full hierarchical name
                "leaf_header": leaf_name,  # Just the final level
                "header_parts": header_parts,  # Individual parts for flexibility
            })

        return column_headers

    def _analyze_table_structure(
        self,
        sheet: Worksheet,
        header_row: int,
        header_cells: List[tuple],
        end_row: Optional[int] = None,
    ) -> Optional[Dict[str, Any]]:
        """
        Analyze a potential table structure starting from a header row.

        Note: This method is typically overridden by TemplateAnalyzer for merged cell support.

        Args:
            sheet: Worksheet
            header_row: Row index of headers
            header_cells: List of (col_idx, header_text) tuples
            end_row: Optional end row of table data

        Returns:
            Table metadata dict or None if not a valid table
        """
        # Base implementation (usually overridden by TemplateAnalyzer)
        # Returns None since we can't properly handle merged cells without it
        return None

    def _deduplicate_tables(self, tables: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
        """
        Remove overlapping/duplicate table detections.

        Args:
            tables: List of detected tables

        Returns:
            Deduplicated list
        """
        if len(tables) <= 1:
            return tables

        # Sort by row, then by column
        sorted_tables = sorted(tables, key=lambda t: (t["start_row"], t["start_col"]))

        deduped = []
        for table in sorted_tables:
            # Check if this table overlaps with any already added
            overlaps = False
            for existing in deduped:
                if self._tables_overlap(table, existing):
                    overlaps = True
                    break

            if not overlaps:
                deduped.append(table)

        return deduped

    def _tables_overlap(self, table1: Dict, table2: Dict) -> bool:
        """Check if two tables overlap in row/column space."""
        # Check row overlap
        table1_end_row = table1.get("end_row")
        if table1_end_row is None:
            table1_end_row = max(table1.get("data_rows", []), default=table1.get("start_row"))

        table2_end_row = table2.get("end_row")
        if table2_end_row is None:
            table2_end_row = max(table2.get("data_rows", []), default=table2.get("start_row"))

        row_overlap = not (
            table1["start_row"] > table2_end_row or
            table2["start_row"] > table1_end_row
        )

        # Check column overlap
        col_overlap = not (
            table1["start_col"] > table2["end_col"] or
            table2["start_col"] > table1["end_col"]
        )

        return row_overlap and col_overlap
